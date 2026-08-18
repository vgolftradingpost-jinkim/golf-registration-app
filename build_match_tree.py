# -*- coding: utf-8 -*-
"""
build_match_tree.py — 매칭 엑셀 → 자동완성 JSON 2종 변환 (+ 월 1회 수작업 반영)
================================================================
입력 : data/00 matching data.xlsx  (시트 'final', 컬럼: TYPE / BRAND / MODEL / SHAFT)
       data/incoming/*.xlsx|*.csv  (선택) 손으로 채워 넣은 신규 데이터
출력 : app/data/match_tree.json    (TYPE > BRAND > MODEL:count 계층)
       app/data/shaft_index.json   (byModel / byBrand / byType — SHAFT 폴백용)

실행 : (프로젝트 폴더 03 registration_app 에서)
       py build_match_tree.py
       py build_match_tree.py --dry-run                 ← 흡수 없이 미리보기
       py build_match_tree.py "data/00 matching data.xlsx"

월 1회 운영 흐름 (수작업):
       data/incoming/_template.xlsx 를 복사 → 새 이름으로 저장
       → 신규 TYPE/BRAND/MODEL/SHAFT 를 손으로 채움
       → data/incoming/ 에 두고  py build_match_tree.py --dry-run  으로 확인
       → py build_match_tree.py   (마스터 흡수 + JSON 재생성)
       → push.bat                 (sw.js CACHE_VERSION 자동 갱신 → 폰 PWA 수신)

규칙:
  · 파일명이 '_' 로 시작하면 건너뜀 → '_template.xlsx' 는 절대 흡수되지 않는다.
  · 같은 파일을 두 번 넣어도 중복 집계되지 않는다(내용 해시 대장으로 판별).
  · 흡수한 파일은 data/incoming/done/ 으로 이동.
  · 마스터는 덮어쓰기 전 data/backup/ 에 자동 백업.
================================================================
"""
import sys, os, csv, json, glob, shutil, hashlib
from datetime import datetime
from collections import defaultdict, Counter

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 필요: pip install openpyxl")

# ---- 경로 ----
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(SCRIPT_DIR, "data", "00 matching data.xlsx")

ARGS = [a for a in sys.argv[1:] if not a.startswith("-")]
FLAGS = {a for a in sys.argv[1:] if a.startswith("-")}
DRY_RUN = "--dry-run" in FLAGS or "-n" in FLAGS

XLSX_PATH = ARGS[0] if ARGS else DEFAULT_XLSX
DATA_DIR = os.path.dirname(os.path.abspath(XLSX_PATH))
INCOMING_DIR = os.path.join(DATA_DIR, "incoming")
DONE_DIR = os.path.join(INCOMING_DIR, "done")
BACKUP_DIR = os.path.join(DATA_DIR, "backup")
LEDGER_PATH = os.path.join(DONE_DIR, "_processed.json")

OUT_DIR = os.path.join(SCRIPT_DIR, "app", "data")
TREE_PATH = os.path.join(OUT_DIR, "match_tree.json")
SHAFT_PATH = os.path.join(OUT_DIR, "shaft_index.json")

HEADERS = ["TYPE", "BRAND", "MODEL", "SHAFT"]

# TYPE 표준화 (소문자 흔들림 보정 — 'wood' 1건 등)
TYPE_NORMALIZE = {
    "driver": "Driver", "wood": "Wood", "hybrid": "Hybrid",
    "iron set": "Iron Set", "iron": "Iron Set",
    "wedge": "Wedge", "putter": "Putter", "etc": "Etc",
}


def norm_type(t):
    if not t:
        return ""
    key = str(t).strip().lower()
    return TYPE_NORMALIZE.get(key, str(t).strip())


def clean(x):
    return str(x).strip() if x is not None else ""


def is_header(row):
    """첫 행이 헤더인지 판정 (TYPE 으로 시작하면 헤더)"""
    return bool(row) and clean(row[0]).upper() == "TYPE"


def norm_row(vals):
    """길이 무관 입력 → (TYPE, BRAND, MODEL, SHAFT). 필수값 없으면 None"""
    v = list(vals) + [None] * (4 - len(vals))
    t, b, m, s = norm_type(v[0]), clean(v[1]), clean(v[2]), clean(v[3])
    if not (t and b and m):
        return None
    return (t, b, m, s)


def file_hash(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def load_ledger():
    if not os.path.exists(LEDGER_PATH):
        return {}
    try:
        with open(LEDGER_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"  [!] 대장 읽기 실패, 새로 시작합니다 — {e}")
        return {}


def save_ledger(led):
    os.makedirs(DONE_DIR, exist_ok=True)
    with open(LEDGER_PATH, "w", encoding="utf-8") as f:
        json.dump(led, f, ensure_ascii=False, indent=1)


# ================================================================
#  incoming 읽기
# ================================================================
def read_incoming_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["final"] if "final" in wb.sheetnames else wb.worksheets[0]
    out, first = [], True
    for row in ws.iter_rows(min_row=1, max_col=4, values_only=True):
        if first:
            first = False
            if is_header(row):
                continue
        r = norm_row(row)
        if r:
            out.append(r)
    wb.close()
    return out


def read_incoming_csv(path):
    out = []
    with open(path, newline="", encoding="utf-8-sig") as f:   # 엑셀 CSV = BOM 포함
        for i, row in enumerate(csv.reader(f)):
            if i == 0 and is_header(row):
                continue
            r = norm_row(row)
            if r:
                out.append(r)
    return out


def collect_incoming():
    """[(파일경로, 해시, [행...]), ...] — done/ 하위와 '_'/'~$' 로 시작하는 파일 제외"""
    if not os.path.isdir(INCOMING_DIR):
        return []
    files = sorted(
        glob.glob(os.path.join(INCOMING_DIR, "*.xlsx"))
        + glob.glob(os.path.join(INCOMING_DIR, "*.csv"))
    )
    batches = []
    for p in files:
        name = os.path.basename(p)
        if name.startswith("_"):        # _template.xlsx 등 — 양식 파일은 흡수 대상 아님
            continue
        if name.startswith("~$"):       # 엑셀 임시 파일
            continue
        try:
            rows = (read_incoming_xlsx(p) if p.lower().endswith(".xlsx")
                    else read_incoming_csv(p))
        except Exception as e:
            print(f"  [!] 읽기 실패, 건너뜀: {name} — {e}")
            continue
        batches.append((p, file_hash(p), rows))
    return batches


# ================================================================
#  JSON 빌드
# ================================================================
def build_json(rows):
    # ---- 1) match_tree: TYPE > BRAND > MODEL:count ----
    tree = defaultdict(lambda: defaultdict(Counter))
    for t, b, m, s in rows:
        if t and b and m:
            tree[t][b][m] += 1

    tree_out = {}
    for t in sorted(tree):
        tree_out[t] = {}
        # 브랜드는 등장 빈도순 정렬
        brand_freq = Counter()
        for b in tree[t]:
            brand_freq[b] = sum(tree[t][b].values())
        for b, _ in brand_freq.most_common():
            # 모델 빈도순 dict (삽입순 = 빈도순)
            tree_out[t][b] = dict(tree[t][b].most_common())

    # ---- 2) shaft_index: byModel / byBrand / byType ----
    by_model = defaultdict(Counter)   # "BRAND||MODEL" -> shaft:count
    by_brand = defaultdict(Counter)   # "BRAND"        -> shaft:count
    by_type = defaultdict(Counter)    # "TYPE"         -> shaft:count
    for t, b, m, s in rows:
        if not s:
            continue
        if b and m:
            by_model[f"{b}||{m}"][s] += 1
        if b:
            by_brand[b][s] += 1
        if t:
            by_type[t][s] += 1

    shaft_out = {
        "byModel": {k: dict(v.most_common()) for k, v in by_model.items()},
        "byBrand": {k: dict(v.most_common()) for k, v in by_brand.items()},
        # byType 는 폴백 최종 단계 — 상위 12개만 저장(앱에서 8개 사용, 여유분)
        "byType":  {k: dict(v.most_common(12)) for k, v in by_type.items()},
    }
    return tree_out, shaft_out


def main():
    if not os.path.exists(XLSX_PATH):
        sys.exit(f"엑셀 파일 없음: {XLSX_PATH}")

    # ---- 마스터 로드 ----
    wb = openpyxl.load_workbook(XLSX_PATH)          # data_only=False (수식 없음 확인됨)
    ws = wb["final"] if "final" in wb.sheetnames else wb.worksheets[0]

    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 5)]
        if all(v is None for v in vals):
            continue
        nr = norm_row(vals)
        if nr:
            rows.append(nr)
    existing = set(rows)
    print(f"기준 데이터: {len(rows)}건 (고유 조합 {len(existing)}개)")

    # ---- incoming 흡수 ----
    ledger = load_ledger()
    batches = collect_incoming()
    new_rows, used_files, dup_files = [], [], 0

    for path, digest, brows in batches:
        name = os.path.basename(path)
        if digest in ledger:
            prev = ledger[digest]
            print(f"  incoming: {name} — 이미 반영된 파일 "
                  f"(원본 '{prev.get('file')}', {prev.get('at')}) → 건너뜀")
            dup_files += 1
            used_files.append((path, digest, name, 0))
            continue
        already = sum(1 for r in brows if r in existing)
        print(f"  incoming: {name} — 읽음 {len(brows)}행"
              + (f" (그중 {already}행은 마스터에 이미 있는 조합 — 빈도로 반영됨)" if already else ""))
        new_rows.extend(brows)
        used_files.append((path, digest, name, len(brows)))

    if not batches:
        print("  incoming 없음 — 기준 데이터로만 빌드")
        print(f"  (신규 반영하려면 {os.path.join('data', 'incoming', '_template.xlsx')} 를 "
              f"복사해 채운 뒤 data/incoming/ 에 두세요)")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if new_rows and not DRY_RUN:
        # 마스터 백업
        os.makedirs(BACKUP_DIR, exist_ok=True)
        base = os.path.splitext(os.path.basename(XLSX_PATH))[0]
        bak = os.path.join(BACKUP_DIR, f"{base}_{stamp}.xlsx")
        shutil.copy2(XLSX_PATH, bak)
        print(f"  백업: {bak}")

        for nr in new_rows:
            ws.append(list(nr))
        wb.save(XLSX_PATH)
        print(f"  마스터 흡수: +{len(new_rows)}행 → {XLSX_PATH}")
    elif new_rows and DRY_RUN:
        print(f"  [dry-run] 흡수 예정 {len(new_rows)}행 — 마스터/파일 변경 없음")

    # 읽은 파일은 (전부 중복이어도) done/ 으로 이동하고 대장에 기록
    if used_files and not DRY_RUN:
        os.makedirs(DONE_DIR, exist_ok=True)
        for path, digest, name, cnt in used_files:
            dest = os.path.join(DONE_DIR, name)
            if os.path.exists(dest):
                root, ext = os.path.splitext(name)
                dest = os.path.join(DONE_DIR, f"{root}_{stamp}{ext}")
            shutil.move(path, dest)
            ledger.setdefault(digest, {
                "file": name,
                "rows": cnt,
                "at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            })
        save_ledger(ledger)
        print(f"  처리 완료 파일 {len(used_files)}개 → {DONE_DIR}"
              + (f" (이미 반영된 파일 {dup_files}개 포함)" if dup_files else ""))

    all_rows = rows + new_rows
    print(f"빌드 대상: {len(all_rows)}건")

    tree_out, shaft_out = build_json(all_rows)

    # ---- 저장 ----
    if DRY_RUN:
        print("[dry-run] JSON 미기록. 종료.")
        return

    os.makedirs(OUT_DIR, exist_ok=True)
    with open(TREE_PATH, "w", encoding="utf-8") as f:
        json.dump(tree_out, f, ensure_ascii=False, separators=(",", ":"))
    with open(SHAFT_PATH, "w", encoding="utf-8") as f:
        json.dump(shaft_out, f, ensure_ascii=False, separators=(",", ":"))

    tree_kb = os.path.getsize(TREE_PATH) // 1024
    shaft_kb = os.path.getsize(SHAFT_PATH) // 1024
    print(f"생성: {TREE_PATH}  ({tree_kb} KB)")
    print(f"  TYPE {len(tree_out)}종, BRAND 조합 "
          f"{sum(len(v) for v in tree_out.values())}개")
    print(f"생성: {SHAFT_PATH}  ({shaft_kb} KB)")
    print(f"  byModel {len(shaft_out['byModel'])} / "
          f"byBrand {len(shaft_out['byBrand'])} / "
          f"byType {len(shaft_out['byType'])}")
    print("완료.")


if __name__ == "__main__":
    main()
